# ABOUTME: Unit tests for excel_export.py module.
# ABOUTME: Tests Excel workbook creation and filename generation.

import re
import sys
from io import BytesIO
from pathlib import Path

import pytest
from openpyxl import load_workbook

# Add src to path for imports
sys.path.insert(0, str(Path(__file__).parent.parent / "src"))

from excel_export import create_testbook_excel, generate_filename, COLUMNS


class TestCreateTestbookExcel:
    """Tests for Excel workbook creation."""

    def test_returns_bytesio(self):
        result = create_testbook_excel([])
        assert isinstance(result, BytesIO)

    def test_empty_test_cases(self):
        result = create_testbook_excel([])
        wb = load_workbook(result)
        ws = wb.active
        assert ws.title == "Test Cases"
        assert ws.max_row == 1  # Only header row

    def test_header_row_present(self):
        result = create_testbook_excel([])
        wb = load_workbook(result)
        ws = wb.active

        expected_headers = [col[1] for col in COLUMNS]
        actual_headers = [ws.cell(row=1, column=i + 1).value for i in range(len(COLUMNS))]
        assert actual_headers == expected_headers

    def test_single_test_case(self):
        test_cases = [{
            "test_case_id": "TC01",
            "test_case_name": "Verify Login",
            "precondition": "User is on login page",
            "steps": "1. Enter username\n2. Enter password\n3. Click login",
            "expected_result": "User is logged in",
        }]
        result = create_testbook_excel(test_cases)
        wb = load_workbook(result)
        ws = wb.active

        assert ws.max_row == 2  # Header + 1 data row
        assert ws.cell(row=2, column=1).value == "TC01"
        assert ws.cell(row=2, column=2).value == "Verify Login"

    def test_multiple_test_cases(self):
        test_cases = [
            {"test_case_id": "TC01", "test_case_name": "First"},
            {"test_case_id": "TC02", "test_case_name": "Second"},
            {"test_case_id": "TC03", "test_case_name": "Third"},
        ]
        result = create_testbook_excel(test_cases)
        wb = load_workbook(result)
        ws = wb.active

        assert ws.max_row == 4  # Header + 3 data rows

    def test_missing_fields_handled(self):
        test_cases = [{"test_case_id": "TC01"}]  # Missing most fields
        result = create_testbook_excel(test_cases)
        wb = load_workbook(result)
        ws = wb.active

        assert ws.cell(row=2, column=1).value == "TC01"
        # openpyxl returns None for empty cells
        assert ws.cell(row=2, column=2).value is None

    def test_freeze_panes_set(self):
        result = create_testbook_excel([])
        wb = load_workbook(result)
        ws = wb.active
        assert ws.freeze_panes == "A2"


class TestGenerateFilename:
    """Tests for filename generation."""

    def test_returns_string(self):
        result = generate_filename()
        assert isinstance(result, str)

    def test_starts_with_testbook(self):
        result = generate_filename()
        assert result.startswith("testbook_")

    def test_ends_with_xlsx(self):
        result = generate_filename()
        assert result.endswith(".xlsx")

    def test_contains_timestamp_pattern(self):
        result = generate_filename()
        # Pattern: testbook_YYYYMMDD_HHMMSS.xlsx
        pattern = r"testbook_\d{8}_\d{6}\.xlsx"
        assert re.match(pattern, result)
