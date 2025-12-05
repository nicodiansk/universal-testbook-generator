# ABOUTME: Excel export functionality for generated test cases.
# ABOUTME: Creates styled Excel workbooks matching the standard testbook format.

from datetime import datetime
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# Column configuration
COLUMNS = [
    ("test_case_id", "Test Case ID", 12),
    ("test_case_name", "Test Case Name", 40),
    ("precondition", "Precondition", 50),
    ("steps", "Steps", 60),
    ("expected_result", "Expected Result", 50),
    ("Result", "Result", 12),
    ("Note", "Note", 30),
]

# Styles
HEADER_FILL = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF")
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)
WRAP_ALIGNMENT = Alignment(wrap_text=True, vertical="top")


def create_testbook_excel(test_cases: list[dict]) -> BytesIO:
    """Create an Excel workbook from test cases and return as BytesIO."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Test Cases"

    # Write header row
    for col_idx, (_, header, width) in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.border = THIN_BORDER
        cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # Write data rows
    for row_idx, tc in enumerate(test_cases, start=2):
        for col_idx, (key, _, _) in enumerate(COLUMNS, start=1):
            value = tc.get(key, "") if key in tc else ""
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = THIN_BORDER
            cell.alignment = WRAP_ALIGNMENT

    # Freeze header row
    ws.freeze_panes = "A2"

    # Save to BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def generate_filename() -> str:
    """Generate timestamped filename for the testbook."""
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    return f"testbook_{timestamp}.xlsx"
