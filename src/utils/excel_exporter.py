"""
Excel export functionality for the Universal Testbook Generator.
Creates professional Excel testbooks ready for QA team execution.
"""

import io
from datetime import datetime
from typing import List, Dict, Any, Optional

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from loguru import logger

from ..models.testbook_models import Testbook, TestProcedure, TestStep
from ..models.document_models import StructuredDocument

class ExcelExporter:
    """Optimized Excel export for QA teams."""
    
    def __init__(self):
        self.header_font = Font(bold=True, color="FFFFFF")
        self.header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        self.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        self.center_alignment = Alignment(horizontal='center', vertical='center')
        self.wrap_alignment = Alignment(wrap_text=True, vertical='top')
    
    def export_testbook(self, testbook: Testbook, template: str = "comprehensive") -> io.BytesIO:
        """
        Export testbook to Excel format.
        
        Args:
            testbook: The testbook to export
            template: Export template type
            
        Returns:
            BytesIO containing the Excel file
        """
        try:
            logger.info(f"Exporting testbook '{testbook.name}' to Excel")
            
            workbook = Workbook()
            workbook.remove(workbook.active)  # Remove default sheet
            
            # Create sheets based on template
            if template == "comprehensive":
                self._create_summary_sheet(workbook, testbook)
                self._create_test_execution_sheet(workbook, testbook)
                self._create_traceability_matrix(workbook, testbook)
                self._create_category_sheets(workbook, testbook)
            else:
                self._create_test_execution_sheet(workbook, testbook)
            
            # Save to BytesIO
            output = io.BytesIO()
            workbook.save(output)
            output.seek(0)
            
            logger.info(f"Successfully exported testbook with {len(workbook.sheetnames)} sheets")
            return output
            
        except Exception as e:
            logger.error(f"Error exporting testbook to Excel: {str(e)}")
            raise
    
    def _create_summary_sheet(self, workbook: Workbook, testbook: Testbook) -> None:
        """Create summary dashboard sheet."""
        sheet = workbook.create_sheet("Summary Dashboard")
        
        # Title
        sheet['A1'] = testbook.name
        sheet['A1'].font = Font(size=16, bold=True)
        sheet.merge_cells('A1:E1')
        
        # Basic information
        info_data = [
            ["Description", testbook.description],
            ["Version", testbook.version],
            ["Created", testbook.created_at.strftime("%Y-%m-%d %H:%M")],
            ["Total Procedures", len(testbook.get_all_procedures())],
            ["Estimated Time", f"{testbook.total_estimated_time} minutes"],
            ["", ""],
            ["Category", "Count"],
            ["Functional Tests", len(testbook.functional_procedures)],
            ["Security Tests", len(testbook.security_procedures)],
            ["Performance Tests", len(testbook.performance_procedures)],
            ["Integration Tests", len(testbook.integration_procedures)]
        ]
        
        for i, (label, value) in enumerate(info_data, start=3):
            sheet[f'A{i}'] = label
            sheet[f'B{i}'] = value
            if label == "Category":  # Header row
                sheet[f'A{i}'].font = self.header_font
                sheet[f'B{i}'].font = self.header_font
                sheet[f'A{i}'].fill = self.header_fill
                sheet[f'B{i}'].fill = self.header_fill
        
        # Auto-adjust column widths
        for column in sheet.columns:
            max_length = 0
            column = [cell for cell in column]
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            sheet.column_dimensions[column[0].column_letter].width = adjusted_width
    
    def _create_test_execution_sheet(self, workbook: Workbook, testbook: Testbook) -> None:
        """Create main test execution sheet."""
        sheet = workbook.create_sheet("Test Execution")
        
        # Headers
        headers = [
            "Test ID", "Test Title", "Category", "Priority", "Description",
            "Preconditions", "Test Steps", "Expected Results", "Evidence Required",
            "Estimated Time (min)", "Actual Result", "Status", "Executed By",
            "Execution Date", "Comments", "Screenshots"
        ]
        
        for col, header in enumerate(headers, start=1):
            cell = sheet.cell(row=1, column=col, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = self.center_alignment
            cell.border = self.border
        
        # Add test procedures
        row = 2
        for procedure in testbook.get_all_procedures():
            # Format test steps
            test_steps_text = "\n".join([
                f"{step.step_number}. {step.action}\n   Expected: {step.expected_behavior}"
                for step in procedure.test_steps
            ])
            
            # Format expected results
            expected_results_text = "\n".join([
                f"• {result.description}" for result in procedure.expected_results
            ])
            
            # Format evidence requirements
            evidence_text = "\n".join([
                f"• {ev.description} ({'Required' if ev.mandatory else 'Optional'})"
                for ev in procedure.evidence_requirements
            ])
            
            # Format preconditions
            preconditions_text = "\n".join([f"• {pc}" for pc in procedure.preconditions])
            
            data = [
                procedure.id,
                procedure.title,
                procedure.category.value.title(),
                procedure.priority.value.title(),
                procedure.description,
                preconditions_text,
                test_steps_text,
                expected_results_text,
                evidence_text,
                procedure.estimated_duration or 0,
                "",  # Actual Result
                "",  # Status
                "",  # Executed By
                "",  # Execution Date
                "",  # Comments
                ""   # Screenshots
            ]
            
            for col, value in enumerate(data, start=1):
                cell = sheet.cell(row=row, column=col, value=value)
                cell.border = self.border
                if col in [6, 7, 8, 9]:  # Wrap text for longer content
                    cell.alignment = self.wrap_alignment
            
            row += 1
        
        # Set column widths
        column_widths = [15, 40, 12, 10, 50, 30, 60, 40, 30, 12, 30, 10, 20, 15, 30, 20]
        for i, width in enumerate(column_widths, start=1):
            sheet.column_dimensions[sheet.cell(row=1, column=i).column_letter].width = width
        
        # Set row heights for better readability
        for row in range(2, sheet.max_row + 1):
            sheet.row_dimensions[row].height = 60
    
    def _create_traceability_matrix(self, workbook: Workbook, testbook: Testbook) -> None:
        """Create requirements traceability matrix."""
        sheet = workbook.create_sheet("Traceability Matrix")
        
        # Headers
        headers = ["Test ID", "Test Title", "Category", "Source Requirements", "Source Features", "Coverage Score"]
        
        for col, header in enumerate(headers, start=1):
            cell = sheet.cell(row=1, column=col, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = self.center_alignment
            cell.border = self.border
        
        # Add traceability data
        row = 2
        for procedure in testbook.get_all_procedures():
            requirements_text = ", ".join(procedure.source_requirements) if procedure.source_requirements else "None"
            features_text = ", ".join(procedure.source_features) if procedure.source_features else "None"
            
            # Simple coverage score calculation
            coverage_score = "High" if procedure.source_requirements else "Low"
            
            data = [
                procedure.id,
                procedure.title,
                procedure.category.value.title(),
                requirements_text,
                features_text,
                coverage_score
            ]
            
            for col, value in enumerate(data, start=1):
                cell = sheet.cell(row=row, column=col, value=value)
                cell.border = self.border
            
            row += 1
        
        # Auto-adjust column widths
        for column in sheet.columns:
            max_length = 0
            column = [cell for cell in column]
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            sheet.column_dimensions[column[0].column_letter].width = adjusted_width
    
    def _create_category_sheets(self, workbook: Workbook, testbook: Testbook) -> None:
        """Create separate sheets for each test category."""
        categories = [
            ("Functional Tests", testbook.functional_procedures),
            ("Security Tests", testbook.security_procedures),
            ("Performance Tests", testbook.performance_procedures),
            ("Integration Tests", testbook.integration_procedures)
        ]
        
        for category_name, procedures in categories:
            if not procedures:
                continue
                
            sheet = workbook.create_sheet(category_name)
            
            # Headers
            headers = ["Test ID", "Title", "Priority", "Steps", "Expected Results", "Status", "Notes"]
            
            for col, header in enumerate(headers, start=1):
                cell = sheet.cell(row=1, column=col, value=header)
                cell.font = self.header_font
                cell.fill = self.header_fill
                cell.alignment = self.center_alignment
                cell.border = self.border
            
            # Add procedures
            row = 2
            for procedure in procedures:
                steps_text = "\n".join([
                    f"{step.step_number}. {step.action}"
                    for step in procedure.test_steps
                ])
                
                results_text = "\n".join([
                    f"• {result.description}" for result in procedure.expected_results
                ])
                
                data = [
                    procedure.id,
                    procedure.title,
                    procedure.priority.value.title(),
                    steps_text,
                    results_text,
                    "",  # Status
                    ""   # Notes
                ]
                
                for col, value in enumerate(data, start=1):
                    cell = sheet.cell(row=row, column=col, value=value)
                    cell.border = self.border
                    if col in [4, 5]:  # Wrap text for steps and results
                        cell.alignment = self.wrap_alignment
                
                row += 1
            
            # Set column widths
            column_widths = [15, 50, 12, 60, 40, 12, 30]
            for i, width in enumerate(column_widths, start=1):
                sheet.column_dimensions[sheet.cell(row=1, column=i).column_letter].width = width
    
    def create_test_execution_sheet(self, procedures: List[TestProcedure]) -> pd.DataFrame:
        """
        Create a pandas DataFrame for test execution tracking.
        
        Args:
            procedures: List of test procedures
            
        Returns:
            DataFrame formatted for test execution
        """
        data = []
        
        for procedure in procedures:
            # Format test steps as numbered list
            steps_text = "; ".join([
                f"{step.step_number}. {step.action}" 
                for step in procedure.test_steps
            ])
            
            # Format expected results
            results_text = "; ".join([
                result.description for result in procedure.expected_results
            ])
            
            data.append({
                "Test_ID": procedure.id,
                "Title": procedure.title,
                "Category": procedure.category.value.title(),
                "Priority": procedure.priority.value.title(),
                "Description": procedure.description,
                "Preconditions": "; ".join(procedure.preconditions),
                "Test_Steps": steps_text,
                "Expected_Results": results_text,
                "Estimated_Time_Min": procedure.estimated_duration or 0,
                "Actual_Result": "",
                "Status": "",
                "Executed_By": "",
                "Execution_Date": "",
                "Comments": "",
                "Evidence_Links": ""
            })
        
        return pd.DataFrame(data)
    
    def add_evidence_collection_columns(self, df: pd.DataFrame) -> pd.DataFrame:
        """
        Add evidence collection columns to a test execution DataFrame.
        
        Args:
            df: Test execution DataFrame
            
        Returns:
            DataFrame with evidence collection columns
        """
        evidence_columns = [
            "Screenshot_1", "Screenshot_2", "Screenshot_3",
            "Log_Files", "Video_Evidence", "Additional_Evidence"
        ]
        
        for col in evidence_columns:
            df[col] = ""
        
        return df
    
    def create_traceability_matrix(
        self, 
        testbook: Testbook, 
        document: Optional[StructuredDocument] = None
    ) -> pd.DataFrame:
        """
        Create a requirements traceability matrix.
        
        Args:
            testbook: The testbook
            document: Optional source document for requirements
            
        Returns:
            DataFrame with traceability information
        """
        data = []
        
        # Get all procedures
        all_procedures = testbook.get_all_procedures()
        
        for procedure in all_procedures:
            data.append({
                "Test_ID": procedure.id,
                "Test_Title": procedure.title,
                "Category": procedure.category.value.title(),
                "Priority": procedure.priority.value.title(),
                "Source_Requirements": ", ".join(procedure.source_requirements) if procedure.source_requirements else "None",
                "Source_Features": ", ".join(procedure.source_features) if procedure.source_features else "None",
                "Tags": ", ".join(procedure.tags) if procedure.tags else "None",
                "Coverage_Status": "Covered" if procedure.source_requirements else "No Requirement Link"
            })
        
        return pd.DataFrame(data)
    
    def format_for_test_management_tools(self, testbook: Testbook, tool: str = "generic") -> io.BytesIO:
        """
        Format testbook for specific test management tools.
        
        Args:
            testbook: The testbook to format
            tool: Target tool (generic, jira, testlink, etc.)
            
        Returns:
            BytesIO with formatted content
        """
        if tool.lower() == "jira":
            return self._format_for_jira(testbook)
        elif tool.lower() == "testlink":
            return self._format_for_testlink(testbook)
        else:
            return self.export_testbook(testbook, template="comprehensive")
    
    def _format_for_jira(self, testbook: Testbook) -> io.BytesIO:
        """Format for JIRA import."""
        data = []
        
        for procedure in testbook.get_all_procedures():
            # Format for JIRA CSV import
            steps_text = "\n".join([
                f"Step {step.step_number}: {step.action}\nExpected: {step.expected_behavior}"
                for step in procedure.test_steps
            ])
            
            data.append({
                "Summary": procedure.title,
                "Description": procedure.description,
                "Test Type": procedure.category.value.title(),
                "Priority": procedure.priority.value.title(),
                "Test Steps": steps_text,
                "Labels": ",".join(procedure.tags) if procedure.tags else ""
            })
        
        df = pd.DataFrame(data)
        output = io.BytesIO()
        df.to_csv(output, index=False, encoding='utf-8')
        output.seek(0)
        return output
    
    def _format_for_testlink(self, testbook: Testbook) -> io.BytesIO:
        """Format for TestLink import."""
        # TestLink XML format would be more complex
        # For now, return standard Excel format
        return self.export_testbook(testbook, template="comprehensive")